import argparse
import os
import zipfile
import tempfile
from openpyxl import load_workbook
from datetime import datetime
import re
from dateutil.relativedelta import relativedelta

from anonymise import StopProcessing
from anonymise import process_files


def row_iterator(filename):
    """Provide a way to iterate over every row as a dict in the given file
    """
    # The default na_values will convert sodium (NA) into NaN!
    cols = [
        "episide",
        "date_collected",
        "time_collected",
        "date_received",
        "time_received",
        "dept",
        "result",
        "stat",
        "test_code",
        "dob",
        "sex",
        "requester",
        "source",
        "patient_numer",
        "patient_category",
    ]
    wb = load_workbook(filename)
    ws = wb.active
    for row in ws.iter_rows():
        yield dict(zip(cols, [str(x.value) for x in row]))


def _date_string_to_past_datetime(date_str):
    try:
        d = datetime.strptime(date_str, "%d/%m/%y")
    except ValueError:
        d = datetime.strptime(date_str, "%d/%m/%Y")
    except ValueError:
        d = datetime.strptime(date_str, "%d %b %Y")
    except ValueError:
        d = datetime.strptime(date_str, "%d %b %y")
    if d > datetime.now():
        d -= relativedelta(years=100)
    return d


def drop_unwanted_data(row):
    """Drop any rows of test data, obviously corrupted data, or otherwise
        unusable data (e.g. no information about the patient's age or the
        practice)
        """
    # Dropping %s null patients (see #62) <- XXX
    if not row["dob"]:
        raise StopProcessing()
    # only GP and A&E
    if row["patient_category"] not in ["GP", "ZE"]:
        raise StopProcessing()


def normalise_data(self):
    """Convert test results to float wherever possible; extract a
    direction if required; set age from DOB; format the date to
    %Y/%m/01.

    Additionally, rename the fields to the standardised list.

    """
    #  Where codes have changed over time, normalise them back to
    #  their current incarnation
    test_code_changes = {
        "AFP3": "AFP2",
        "ACE1": "ACE",
        "AT3S": "AT3",
        "FDP1": "FDP",
        "FPSS": "FPS",
        "INR1": "INR",
        "PT1": "PT",
    }
    self.row["test_code"] = test_code_changes.get(
        self.row["test_code"], self.row["test_code"]
    )
    dob = _date_string_to_past_datetime(self.row["dob"])
    collected = _date_string_to_past_datetime(self.row["date_collected"])

    self.row["age"] = (collected - dob).days / 365
    result = self.row["result"]
    self.row["month"] = collected.strftime("%Y/%m/01")
    direction = None
    try:
        if result.startswith("<"):
            direction = "<"
            result = float(result[1:]) - 0.0000001
        elif result.startswith(">"):
            direction = ">"
            result = float(result[1:]) + 0.0000001
        else:
            result = float(result)
        self.row["test_result"] = result
        self.row["direction"] = direction

        col_mapping = {
            "month": "month",
            "test_code": "test_code",
            "test_result": "test_result",
            "practice_id": "source",
            "age": "age",
            "sex": "sex",
            "direction": "direction",
        }
        mapped = {}
        for k, v in col_mapping.items():
            mapped[k] = self.row[v]
        self.row = mapped
    except ValueError:
        self.log_info("Unparseable result %s", result)
        raise StopProcessing()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate suitably anonymised subset of raw input data"
    )

    parser.add_argument("lab", help="Lab code (used for naming files)")
    parser.add_argument("files", nargs="+", help="Monthly input files")
    parser.add_argument(
        "--multiprocessing", help="Use multiprocessing", action="store_true"
    )
    args = parser.parse_args()
    filters = (drop_unwanted_data,)
    normaliser = normalise_data
    process_files(
        args.lab,
        args.files,
        row_iterator,
        filters,
        normalise_data,
        args.multiprocessing,
    )

import glob
import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta

from lib.anonymise import StopProcessing
from lib.anonymise import log_error

LAB_CODE = "nd"
REFERENCE_RANGES = "north_devon_reference_ranges.csv"
files_path = os.path.join(
    os.environ.get("DATA_BASEDIR", "/home/filr/"), "NorthDevon/*/NDHTSB*"
)
INPUT_FILES = glob.glob(files_path)


def row_iterator(filename):
    """Provide a way to iterate over every row as a dict in the given file
    """
    # The default na_values will convert sodium (NA) into NaN!
    cols = [
        "episide",
        "date_collected",
        "time_collected",
        "date_received",
        "time_received",
        "dept",
        "result",
        "stat",
        "test_code",
        "dob",
        "sex",
        "requester",
        "source",
        "patient_numer",
        "patient_category",
    ]
    try:
        os.symlink(os.path.split(filename)[-1], filename + ".xlsx")
        wb = load_workbook(filename + ".xlsx", read_only=True)
        ws = wb.active
        for row in ws.iter_rows():
            yield dict(zip(cols, [str(x.value) for x in row]))
    finally:
        os.remove(filename + ".xlsx")


def _date_string_to_past_datetime(date_str):
    try:
        d = datetime.strptime(date_str, "%d/%m/%y")
    except ValueError:
        d = datetime.strptime(date_str, "%d/%m/%Y")
    except ValueError:
        d = datetime.strptime(date_str, "%d %b %Y")
    except ValueError:
        d = datetime.strptime(date_str, "%d %b %y")
    if d > datetime.now():
        d -= relativedelta(years=100)
    return d


def drop_unwanted_data(row):
    """Drop any rows of test data, obviously corrupted data, or otherwise
        unusable data (e.g. no information about the patient's age or the
        practice)
        """
    # Dropping %s null patients (see #62)
    if row["dob"] in ["None", ""]:
        raise StopProcessing()
    # only GP and A&E
    if row["patient_category"] not in ["GP", "ZE"]:
        raise StopProcessing()


PRACTICE_MAP = (
    pd.read_csv("north_devon/north_devon_practice_mapping.csv", na_filter=False)
    .set_index("LIMS code")
    .to_dict(orient="index")
)


def normalise_data(row):
    """Convert test results to float wherever possible; extract a
    direction if required; set age from DOB; format the date to
    %Y/%m/01.

    Additionally, rename the fields to the standardised list.

    """
    # Convert local practice ids to ODS code
    practice_id = PRACTICE_MAP.get(row["source"], {"ODS code": ""})["ODS code"]
    if not practice_id:
        raise StopProcessing()
    row["practice_id"] = practice_id

    #  Where codes have changed over time, normalise them back to
    #  their current incarnation
    # XXX now redundant?
    test_code_changes = {
        "AFP3": "AFP2",
        "ACE1": "ACE",
        "AT3S": "AT3",
        "FDP1": "FDP",
        "FPSS": "FPS",
        "INR1": "INR",
        "PT1": "PT",
    }
    row["test_code"] = test_code_changes.get(row["test_code"], row["test_code"])
    try:
        dob = _date_string_to_past_datetime(row["dob"])
    except ValueError:
        log_error(row, "Unable to parse dob")
        raise
    collected = _date_string_to_past_datetime(row["date_collected"])

    row["age"] = (collected - dob).days / 365
    if row["age"] < 18:
        raise StopProcessing()
    result = row["result"]
    row["month"] = collected.strftime("%Y/%m/01")
    direction = None
    try:
        if result.startswith("<"):
            direction = "<"
            result = float(result[1:]) - 0.0000001
        elif result.startswith(">"):
            direction = ">"
            result = float(result[1:]) + 0.0000001
        else:
            result = float(result)
    except ValueError:
        pass
    row["test_result"] = result
    row["direction"] = direction

    col_mapping = {
        "month": "month",
        "test_code": "test_code",
        "test_result": "test_result",
        "practice_id": "practice_id",
        "age": "age",
        "sex": "sex",
        "direction": "direction",
    }
    mapped = {}
    for k, v in col_mapping.items():
        mapped[k] = row[v]
    return mapped

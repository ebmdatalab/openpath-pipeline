import glob
import os
import pandas as pd

from datetime import datetime
from openpyxl import load_workbook

from lib.intermediate_file_processing import StopProcessing
from lib.logger import log_info, log_warning

LAB_CODE = "exeter"
REFERENCE_RANGES = ""

files_path = os.path.join(
    os.environ.get("DATA_BASEDIR", "/home/filr/"), "Exeter/*.xlsx"
)
INPUT_FILES = glob.glob(files_path)

RANGE_CEILING = 99999


# Error codes
WITHIN_RANGE = 0
UNDER_RANGE = -1
OVER_RANGE = 1
ERR_NO_REF_RANGE = 2
ERR_UNPARSEABLE_RESULT = 3
ERR_INVALID_SEX = 4
ERR_INVALID_RANGE_WITH_DIRECTION = 5
ERR_DISCARDED_AGE = 6
ERR_INVALID_REF_RANGE = 7
ERR_NO_TEST_CODE = 8


def row_iterator(filename):
    """Provide a way to iterate over every row as a dict in the given file
    """
    required_cols = [
        "Date_Request_Made",
        "Requesting_Organisation_Code",
        "Requesting_Organisation_Desc",
        "Age_on_Date_Request_Rec'd",
        "Sex",
        "Requested_Test_Code",
        "Test_Performed",
        "Date_Test_Performed",
        "Date_Request_Made",
        "Date_Specimen_Received",
        "Test_Result",
        "Test_Result_Range",
        "Test_Result_Units",
    ]
    wb = load_workbook(filename, read_only=True)
    ws = wb.active
    keys = []
    for row in ws.iter_rows():
        if not keys:
            keys = [x.value for x in row]
            # check every element in required_cols is in keys
            assert set(required_cols).issubset(
                set(keys)
            ), "File at {} must define columns {}, has {}".format(
                filename, required_cols, keys
            )

        else:
            yield dict(zip(keys, [str(x.value) for x in row]))


def drop_unwanted_data(row):
    """Drop any rows of test data, obviously corrupted data, or otherwise
        unusable data (e.g. no information about the patient's age or the
        practice)
        """
    age = row["Age_on_Date_Request_Rec'd"]
    if age[:2] < "18":
        raise StopProcessing()
    if "Hospital" in row["Requesting_Organisation_Desc"]:
        raise StopProcessing()


PRACTICE_MAP = (
    pd.read_csv(os.path.join(os.path.dirname(__file__), "exeter_practices_branch.csv"))
    .set_index("Requesting_Organisation_Code")[["Parent_Requesting_Organisation_Code"]]
    .dropna()
    .to_dict(orient="index")
)


def normalise_data(row):
    """Convert test results to float wherever possible; extract a
    direction if required; set age from DOB; format the date to
    %Y/%m/01.

    Additionally, rename the fields to the standardised list.

    """
    date_fields = [
        "Date_Request_Made",
        "Date_Specimen_Collected",
        "Date_Specimen_Received",
    ]
    order_date = None
    for date_field in date_fields:
        try:
            order_date = datetime.strptime(row[date_field], "%Y-%m-%d 00:00:00")
        except ValueError:
            pass
    if not order_date:
        log_warning(row, "Unparseable date")
        raise StopProcessing()

    row["month"] = order_date.strftime("%Y/%m/01")
    row["dob"] = ""
    row["age"] = ""
    row["sex"] = ""
    row["test_result"] = ""
    row["direction"] = ""
    row["practice_id"] = PRACTICE_MAP.get(
        row["Requesting_Organisation_Code"], row["Requesting_Organisation_Code"]
    )
    col_mapping = {
        "month": "month",
        "test_code": "Test_Performed",
        "test_result": "test_result",
        "practice_id": "practice_id",
        "age": "age",
        "sex": "sex",
        "direction": "direction",
        "provided_result": "Test_Result_Range",
    }
    mapped = {}
    for k, v in col_mapping.items():
        mapped[k] = row[v]
    return mapped


def convert_to_result(row, ranges):
    """Set a value of the `result_category` key, based on existing fields:

    month, test_code, practice_id, age, sex, direction
    """
    if row["provided_result"] == "H":
        return_code = OVER_RANGE
    elif row["provided_result"] == "L":
        return_code = UNDER_RANGE
    elif row["provided_result"] == "N":
        return_code = WITHIN_RANGE
    else:
        return_code = ERR_NO_REF_RANGE
    row["result_category"] = return_code
    return row
